"""
Password Loader Module
Load credentials from Excel files
"""

from pathlib import Path
from typing import List, Dict, Optional, Tuple
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .logger import get_logger


class Credential:
    """Data class for storing credentials"""

    def __init__(self, username: str, password: str, note: str = ""):
        self.username = username.strip() if username else ""
        self.password = password.strip() if password else ""
        self.note = note.strip() if note else ""

    def __repr__(self):
        return f"Credential(username='{self.username}', password='***', note='{self.note}')"

    def is_valid(self) -> bool:
        """Check if credential has both username and password"""
        return bool(self.username and self.password)

    def to_dict(self) -> dict:
        """Convert to dictionary"""
        return {
            "username": self.username,
            "password": self.password,
            "note": self.note
        }


class PasswordLoader:
    """Load credentials from Data Table files"""

    def __init__(self):
        self.logger = get_logger()
        self.credentials: List[Credential] = []

    def load_from_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        username_column: int = 0,
        password_column: int = 1,
        note_column: Optional[int] = None,
        skip_header: bool = True,
        start_row: int = 1
    ) -> List[Credential]:
        """
        Load credentials from Excel file

        Args:
            file_path: Path to Excel file (.xlsx, .xls)
            sheet_name: Sheet name (None for active sheet)
            username_column: Column index for username (0-based)
            password_column: Column index for password (0-based)
            note_column: Column index for notes (0-based, optional)
            skip_header: Skip first row if True
            start_row: Starting row number (1-based)

        Returns:
            List of Credential objects
        """
        try:
            file_path = Path(file_path)

            if not file_path.exists():
                self.logger.error(f"Excel file not found: {file_path}")
                return []

            self.logger.info(f"Loading credentials from: {file_path}")

            # Load workbook
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            # Get worksheet
            if sheet_name:
                if sheet_name not in wb.sheetnames:
                    self.logger.error(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
                    return []
                ws = wb[sheet_name]
            else:
                ws = wb.active

            self.logger.info(f"Reading from sheet: {ws.title}")

            # Read credentials
            credentials = []
            row_start = start_row + (1 if skip_header else 0)

            for row_idx, row in enumerate(ws.iter_rows(min_row=row_start), start=row_start):
                try:
                    # Get username
                    username_cell = row[username_column] if username_column < len(row) else None
                    username = str(username_cell.value) if username_cell and username_cell.value is not None else ""

                    # Get password
                    password_cell = row[password_column] if password_column < len(row) else None
                    password = str(password_cell.value) if password_cell and password_cell.value is not None else ""

                    # Get note (optional)
                    note = ""
                    if note_column is not None and note_column < len(row):
                        note_cell = row[note_column]
                        note = str(note_cell.value) if note_cell and note_cell.value is not None else ""

                    # Create credential
                    if username or password:  # At least one field should have data
                        credential = Credential(username, password, note)
                        if credential.is_valid():
                            credentials.append(credential)
                            self.logger.debug(f"Row {row_idx}: Loaded credential for '{username}'")
                        else:
                            self.logger.warning(f"Row {row_idx}: Invalid credential (missing username or password)")

                except Exception as e:
                    self.logger.warning(f"Row {row_idx}: Error reading credential - {e}")
                    continue

            wb.close()

            self.credentials = credentials
            self.logger.success(f"Loaded {len(credentials)} valid credentials from Excel file")

            return credentials

        except Exception as e:
            self.logger.error(f"Failed to load Excel file: {e}")
            return []

    def load_from_csv(
        self,
        file_path: str,
        username_column: int = 0,
        password_column: int = 1,
        note_column: Optional[int] = None,
        skip_header: bool = True,
        delimiter: str = ','
    ) -> List[Credential]:
        """
        Load credentials from CSV file

        Args:
            file_path: Path to CSV file
            username_column: Column index for username (0-based)
            password_column: Column index for password (0-based)
            note_column: Column index for notes (0-based, optional)
            skip_header: Skip first row if True
            delimiter: CSV delimiter

        Returns:
            List of Credential objects
        """
        try:
            import csv

            file_path = Path(file_path)

            if not file_path.exists():
                self.logger.error(f"CSV file not found: {file_path}")
                return []

            self.logger.info(f"Loading credentials from CSV: {file_path}")

            credentials = []

            with open(file_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.reader(f, delimiter=delimiter)

                if skip_header:
                    next(reader, None)

                for row_idx, row in enumerate(reader, start=1):
                    try:
                        username = row[username_column].strip() if username_column < len(row) else ""
                        password = row[password_column].strip() if password_column < len(row) else ""
                        note = row[note_column].strip() if note_column is not None and note_column < len(row) else ""

                        if username or password:
                            credential = Credential(username, password, note)
                            if credential.is_valid():
                                credentials.append(credential)
                                self.logger.debug(f"Row {row_idx}: Loaded credential for '{username}'")
                            else:
                                self.logger.warning(f"Row {row_idx}: Invalid credential")

                    except Exception as e:
                        self.logger.warning(f"Row {row_idx}: Error reading credential - {e}")
                        continue

            self.credentials = credentials
            self.logger.success(f"Loaded {len(credentials)} valid credentials from CSV file")

            return credentials

        except Exception as e:
            self.logger.error(f"Failed to load CSV file: {e}")
            return []

    def get_credentials(self) -> List[Credential]:
        """Get loaded credentials"""
        return self.credentials

    def get_credential_count(self) -> int:
        """Get number of loaded credentials"""
        return len(self.credentials)

    def clear_credentials(self):
        """Clear loaded credentials"""
        self.credentials.clear()
        self.logger.info("Credentials cleared")

    def get_sheets(self, file_path: str) -> List[str]:
        """
        Get list of sheet names from Excel file

        Args:
            file_path: Path to Excel file

        Returns:
            List of sheet names
        """
        try:
            file_path = Path(file_path)

            if not file_path.exists():
                self.logger.error(f"Excel file not found: {file_path}")
                return []

            wb = openpyxl.load_workbook(file_path, read_only=True)
            sheets = wb.sheetnames
            wb.close()

            self.logger.debug(f"Found {len(sheets)} sheets: {sheets}")
            return sheets

        except Exception as e:
            self.logger.error(f"Failed to read Excel sheets: {e}")
            return []

    def preview_excel(
        self,
        file_path: str,
        sheet_name: Optional[str] = None,
        max_rows: int = 10
    ) -> Tuple[List[str], List[List[str]]]:
        """
        Preview Excel file content

        Args:
            file_path: Path to Excel file
            sheet_name: Sheet name (None for active sheet)
            max_rows: Maximum number of rows to preview

        Returns:
            Tuple of (column_letters, preview_data)
        """
        try:
            file_path = Path(file_path)

            if not file_path.exists():
                self.logger.error(f"Excel file not found: {file_path}")
                return ([], [])

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

            if sheet_name:
                ws = wb[sheet_name]
            else:
                ws = wb.active

            # Get column letters
            columns = []
            if ws.max_column:
                from openpyxl.utils import get_column_letter
                columns = [get_column_letter(i) for i in range(1, min(ws.max_column + 1, 27))]  # Limit to Z

            # Get preview data
            preview_data = []
            for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows, values_only=True)):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                preview_data.append(row_data)

            wb.close()

            self.logger.debug(f"Preview loaded: {len(preview_data)} rows, {len(columns)} columns")
            return (columns, preview_data)

        except Exception as e:
            self.logger.error(f"Failed to preview Excel file: {e}")
            return ([], [])

    def validate_file(self, file_path: str) -> bool:
        """
        Validate if file is a valid Excel file

        Args:
            file_path: Path to file

        Returns:
            True if valid, False otherwise
        """
        try:
            file_path = Path(file_path)

            if not file_path.exists():
                self.logger.error(f"File not found: {file_path}")
                return False

            if file_path.suffix.lower() not in ['.xlsx', '.xls', '.xlsm']:
                self.logger.error(f"Invalid file format: {file_path.suffix}")
                return False

            # Try to open the file
            wb = openpyxl.load_workbook(file_path, read_only=True)
            wb.close()

            self.logger.debug(f"File validation passed: {file_path}")
            return True

        except Exception as e:
            self.logger.error(f"File validation failed: {e}")
            return False
